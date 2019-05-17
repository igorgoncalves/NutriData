import io, re, uuid, os, shutil
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, colors

# Essa classe é responsável por ler uma planilha e transformar em um dicionario (e seu inverso)

class xslxAdapter():
    NUMERO_DE_LINHAS_FIXAS = 3
    NUMERO_DE_COLUNAS_FIXAS = 1
    _workbook = {}

    def ler_planilha_xlsx(self, arquivo):
        self._workbook = {}
        self._workbook = load_workbook(filename=BytesIO(arquivo.read()))
        macroindicador = arquivo.filename

        #leitura do arquivo xlsx
        anos = self._workbook.sheetnames
        leitura = []
                
        for ws in self._workbook:
            sheet = []
            for row in ws.values:
                sheet.append(row)
            leitura.append(sheet)
        
        objeto_transformado = self._organizer_sheet(leitura, "Macroindicador", "Descricao", anos)

        return objeto_transformado

    def _organizer_sheet(self, planilha2, nome_macroindicador, descricao_macroindicador, anos):
        nome_macroindicador = nome_macroindicador
        descricao_macroindicador = descricao_macroindicador
        output = []

        fonte, indicadores, unidade = planilha2[0][0][1:],planilha2[0][1][1:], planilha2[0][2][1:]        

        valores = []
        for pagina in planilha2:
            valores.append(pagina[self.NUMERO_DE_LINHAS_FIXAS:])
            
        macroindicador = {}
        localidades_com_valores = []
        macroindicador['nome'] = nome_macroindicador
        macroindicador['descricao_macroindicador'] = descricao_macroindicador
        macroindicador['fonte'] = fonte[0]
        macroindicador['unidade'] = unidade[0]
        indicadores_list = []
        amostras_ano = len(valores)

        for idx, nome_do_indicador in enumerate(indicadores):
            indicador = {}
            coluna = idx
            indicador['nome'] = nome_do_indicador
            indicador['indicadores_filho'] = []
            amostras = []
            localidades = []
            for ano in range (0, amostras_ano):
                for territorio in valores[ano]:
                        amostra_obj = {}
                        posicao_localidade_arquivo = valores[ano].index(territorio) + self.NUMERO_DE_LINHAS_FIXAS
                        valor_obtido = territorio[self.NUMERO_DE_COLUNAS_FIXAS:]
                        valor_amostral = valor_obtido[coluna]
                        amostra_obj['posicao_localidade_arquivo'] = posicao_localidade_arquivo
                        amostra_obj['ano'] = anos[ano]
                        amostra_obj['valor'] = valor_amostral
                        amostra_obj['coordenada_planilha'] = { 'coluna': coluna + self.NUMERO_DE_COLUNAS_FIXAS , 'linha': posicao_localidade_arquivo, 'tabela': ano }

                        if valor_amostral != '-' and valor_amostral != None and valor_amostral != 0 and valor_amostral != "...":                        
                            amostras.append(amostra_obj)
                            localidades_com_valores.append(posicao_localidade_arquivo)
                            if not posicao_localidade_arquivo in localidades:
                                localidades.append(posicao_localidade_arquivo)
                                
            indicador['amostras'] = amostras
            indicadores_list.append(indicador)

        aux = set(localidades_com_valores)
        localidades_com_valores = list(aux)
        macroindicador['locais_id'] = localidades_com_valores
        macroindicador['posicoes_localidades'] = localidades
        macroindicador['indicadores'] = indicadores_list
        

        return macroindicador
            
    def sheet_to_dict(self):
        anos = ['2017', '2018'] # os anos vem das folhas do excel
        planilha2 = [[('INDICADORES', '1. Cereais e leguminosas', '1.1. Cereais', '1.2. Leguminosas', '2. Hortaliças', '2.1. Hortaliças folhosas e florais', '2.2. Hortaliças frutais', '2.3. Hortaliças tuberosas e outras', '3. Frutas', '3.1. Frutas de clima tropical', '3.2. Frutas de clima temperado', '4. Cocos, castanhas e nozes', '4.1 Coco', '4.2. Castanhas e nozes', '5. Farinhas, féculas e massas', '5.1. Farinhas', '5.2. Féculas', '5.3. Massas', '6. Panificados', '6.1. Pães', '6.2. Bolos', '6.3. Biscoitos, roscas, etc.', '7. Carnes', '7.1. Carnes bovinas de primeira', '7.2. Carnes bovinas de segunda', '7.3. Carnes bovinas outras', '7.4. Carnes suínas com e sem osso', '7.5. Carnes suínas outras', '7.6. Carnes de outros animais', '8. Vísceras', '8.1. Vísceras bovinas', '8.2. Vísceras suínas', '8.3. Outras vísceras', '9. Pescados', '9.1. Pescados de água salgada', '9.2. Pescados de água doce', '9.3. Pescados não especificados', '10. Aves e ovos', '10.1. Aves', '10.2. Ovos', '11. Laticínios', '11.1. Leite e creme de leite', '11.2. Queijos e requeijão', '11.3. Outros laticínios', '12. Açúcares, doces e produtos de confeitaria', '12.1. Açúcares', '12.2. Doces e produtos de confeitaria', '12.3. Outros açúcares, doces e produtos de confeitaria', '13. Sais e condimentos', '13.1. Sais', '13.2. Condimentos', '14. Óleos e gorduras', '14.1. Óleos', '14.2. Gorduras', '15. Bebidas e infusões', '15.1. Bebidas alcoólicas', '15.2. Bebidas não alcoólicas', '15.3. Cafés', '15.4. Chás', '16. Alimentos preparados e misturas industriais', '16.1. Alimentos preparados', '16.2. Misturas industriais', '17. Outros produtos'), 
                    ('UNIDADE', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita'), 
                    ('Fonte', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE'), 
                    ('SERGIPE', 27.36, 15.811, 11.549, 29.841, 2.052, 15.268, 12.521, 29.46, 26.004, 3.456, 0.513, 0.48, 0.033, 31.713, 14.891, 13.004, 3.817, 22.817, 17.873, 0.514, 4.429, 28.272, 5.06, 10.482, 7.561, 1.009, 2.475, 1.686, 1.369, 1.264, 0.064, 0.041, 5.973, 3.312, 1.347, 1.315, 21.213, 17.166, 4.047, 26.129, 21.773, 1.515, 2.841, 18.381, 16.054, 1.897, 0.429, 5.062, 2.594, 2.468, 6.428, 4.84, 1.588, 34.206, 4.491, 27.998, 1.702, 0.016, 1.834, 1.718, 0.116, 0.07),
                    ], 
                    [('INDICADORES', '1. Cereais e leguminosas', '1.1. Cereais', '1.2. Leguminosas', '2. Hortaliças', '2.1. Hortaliças folhosas e florais', '2.2. Hortaliças frutais', '2.3. Hortaliças tuberosas e outras', '3. Frutas', '3.1. Frutas de clima tropical', '3.2. Frutas de clima temperado', '4. Cocos, castanhas e nozes', '4.1 Coco', '4.2. Castanhas e nozes', '5. Farinhas, féculas e massas', '5.1. Farinhas', '5.2. Féculas', '5.3. Massas', '6. Panificados', '6.1. Pães', '6.2. Bolos', '6.3. Biscoitos, roscas, etc.', '7. Carnes', '7.1. Carnes bovinas de primeira', '7.2. Carnes bovinas de segunda', '7.3. Carnes bovinas outras', '7.4. Carnes suínas com e sem osso', '7.5. Carnes suínas outras', '7.6. Carnes de outros animais', '8. Vísceras', '8.1. Vísceras bovinas', '8.2. Vísceras suínas', '8.3. Outras vísceras', '9. Pescados', '9.1. Pescados de água salgada', '9.2. Pescados de água doce', '9.3. Pescados não especificados', '10. Aves e ovos', '10.1. Aves', '10.2. Ovos', '11. Laticínios', '11.1. Leite e creme de leite', '11.2. Queijos e requeijão', '11.3. Outros laticínios', '12. Açúcares, doces e produtos de confeitaria', '12.1. Açúcares', '12.2. Doces e produtos de confeitaria', '12.3. Outros açúcares, doces e produtos de confeitaria', '13. Sais e condimentos', '13.1. Sais', '13.2. Condimentos', '14. Óleos e gorduras', '14.1. Óleos', '14.2. Gorduras', '15. Bebidas e infusões', '15.1. Bebidas alcoólicas', '15.2. Bebidas não alcoólicas', '15.3. Cafés', '15.4. Chás', '16. Alimentos preparados e misturas industriais', '16.1. Alimentos preparados', '16.2. Misturas industriais', '17. Outros produtos'), 
                    ('UNIDADE', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita', 'Kg/ano per capita'), 
                    ('Fonte', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE', 'IBGE'), 
                    ('SERGIPE', 27.36, 15.811, 11.549, 29.841, 2.052, 15.268, 12.521, 29.46, 26.004, 3.456, 0.513, 0.48, 0.033, 31.713, 14.891, 13.004, 3.817, 22.817, 17.873, 0.514, 4.429, 28.272, 5.06, 10.482, 7.561, 1.009, 2.475, 1.686, 1.369, 1.264, 0.064, 0.041, 5.973, 3.312, 1.347, 1.315, 21.213, 17.166, 4.047, 26.129, 21.773, 1.515, 2.841, 18.381, 16.054, 1.897, 0.429, 5.062, 2.594, 2.468, 6.428, 4.84, 1.588, 34.206, 4.491, 27.998, 1.702, 0.016, 1.834, 1.718, 0.116, 0.07),
                    ]
                    ]
        out = [
                {'nome': 'arquivo',
                'descricao': 'descricao',
                'fonte':'fonte',
                'unidade':'unidade',
                'indicadores': [
                    {
                        'nome' : 'indicador1',
                        'indicadores_filho':'indicadores_filho',
                        'amostras' : [
                            {
                                'ano': 'sheets-name1',
                                'valor':' valor',
                                'local_posicao': '1',
                            },
                            {
                                'ano': 'sheets-name2',
                                'valor':' valor',
                                'local_posicao': '1',

                            }
                        ]
                    },
                    {
                        'nome' : 'indicador2',
                        'indicadores_filho':'indicadores_filho',
                        'amostras' : [
                            {
                                'ano': 'sheets-name1',
                                'valor':' valor',
                                'local_posicao': '1',
                            },
                            {
                                'ano': 'sheets-name2',
                                'valor':' valor',
                                'local_posicao': '1',

                            }
                        ]
                    },
                ]}
        ]
        return out     
        
    def gerar_planilha_para_correcao(self, coordenada_celulas_com_problema):        
        
        for coordenada_celula, mensagem in coordenada_celulas_com_problema:
            pos_tabela, pos_coluna, pos_linha = map(int, [coordenada_celula['tabela'],coordenada_celula['coluna'], coordenada_celula['linha']])
            celula = self._workbook._sheets[pos_tabela].cell(row=pos_linha+1, column=pos_coluna+1)
            comentario = Comment(mensagem, "Sistema")
            celula.font = Font(color=colors.RED)
            celula.comment = comentario
        
        directory = '../dist/static/xl/'
        if not os.path.exists(directory):            
            os.makedirs(directory, exist_ok=True)
        else:
            shutil.rmtree(directory)
            os.makedirs(directory)
        
        # name = "output"
        name = uuid.uuid4().hex

        uri = "static/xl/{0}.xlsx".format(name)
        self._workbook.save("../dist/" + uri)

        return uri
                    
        